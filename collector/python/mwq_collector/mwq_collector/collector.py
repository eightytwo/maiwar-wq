import json
import tempfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# Column and row indexes of relevant data
DATE_ROW = 7
FIRST_DATE_COL = "G"
FIRST_LOCATION_ROW = 8
LAST_LOCATION_ROW = 18
LOCATION_COL = "C"


def _cleanse_measurement(measurement: str) -> int:
    """Ensure a measurement taken from a spreadsheet is a number.
    The string 'NT' is used to represent 'Not tested' and some cells
    contain '<1' instead of an integer value.

    :param measurement: The value of a measurement from the spreadsheet.
    :return: The measurement as an integer.
    """
    if isinstance(measurement, int):
        return measurement
    elif measurement.startswith(("<", ">")):
        return int(measurement.replace("<", "").replace(">", "").replace(",", ""))
    else:
        return -1


def _get_dates(sheet: Worksheet) -> list[datetime]:
    """Extract the measurement dates from the spreadsheet.

    :param sheet: The spreadsheet to read the dates from.
    :return: A list of datetime objects.
    """
    start = f"{FIRST_DATE_COL}{DATE_ROW}"
    end = f"{get_column_letter(sheet.max_column)}{DATE_ROW}"
    dates = [d.value for d in sheet[start:end][0] if d.value]
    return dates


def _get_locations(sheet: Worksheet) -> list[str]:
    """Extract the locations from the spreadsheet.

    :param sheet: The spreadsheet to read the locations from.
    :return: A list of location names.
    """
    start = f"{LOCATION_COL}{FIRST_LOCATION_ROW}"
    end = f"{LOCATION_COL}{LAST_LOCATION_ROW}"
    return [location.value.lower().replace(" ", "-") for location, in sheet[start:end]]


def _get_measurements(sheet: Worksheet) -> list[tuple[int, ...]]:
    """Extract the measurements from the spreadsheet.

    :param sheet: The spreadsheet to read the locations from.
    :return: A list of tuples, with each tuple containing measurements for a date.
    """
    col_iter = sheet.iter_cols(
        min_row=FIRST_LOCATION_ROW,
        max_row=LAST_LOCATION_ROW,
        min_col=column_index_from_string(FIRST_DATE_COL),
        max_col=sheet.max_column,
        values_only=True,
    )

    measurements = []

    for col in col_iter:
        if not col[0]:
            continue

        measurements.append(tuple(_cleanse_measurement(value) for value in col))

    return measurements


def _read_sheet(sheet: Worksheet) -> dict[str, dict[str, int]]:
    """Read date, location, and measurement data from a spreadsheet and represent
    this information as a dictionary. Date keys reference dictionaries of location
    and measurement key-value pairs.

    :param sheet: The spreadsheet to read the locations from.
    :return: A nested dictionary representing the dates and locations measurements
     were taken.
    """
    dates = _get_dates(sheet)
    locations = _get_locations(sheet)
    measurements = _get_measurements(sheet)

    results = defaultdict(dict)

    for date_index, date_value in enumerate(dates):
        date_str = date_value.strftime("%Y-%m-%d")

        for location_index, location in enumerate(locations):
            results[date_str][location] = measurements[date_index][location_index]

    return results


def _transform(workbook_path: str) -> dict:
    """Transform an Excel workbook of measurements into JSON.

    :param workbook_path: The full file path to the Excel workbook.
    :return: A dictionary containing the measurements from the Excel workbook.
    """
    wb = load_workbook(filename=workbook_path)
    result_sheets = [sheet for sheet in wb if "results" in sheet.title]

    all_results = {}

    for sheet in result_sheets:
        results = _read_sheet(sheet)
        all_results.update(results)

    return all_results


def transform(workbook: bytes) -> dict:
    """Transform an Excel workbook of measurements into JSON.

    :param workbook: The workbook data.
    :return: A dictionary containing the measurements from the Excel workbook.
    """
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
        tmp.write(workbook)
        return _transform(tmp.name)


if __name__ == "__main__":
    results = _transform("/tmp/mwq_measurements.xlsx")
    Path("/tmp/mwq_measurements.json").write_text(json.dumps(results, sort_keys=True))
