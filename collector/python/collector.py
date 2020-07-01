import json
from collections import defaultdict
from datetime import datetime
from typing import Dict
from typing import List
from typing import Tuple

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# The Excel file to read and the JSON document to write
XLSX_FILE = 'measurements.xlsx'
RESULTS_FILE = 'measurements_py.json'

# Column and row indexes of relevant data
DATE_ROW = 7
FIRST_DATE_COL = 'G'
FIRST_LOCATION_ROW = 8
LAST_LOCATION_ROW = 18
LOCATION_COL = 'C'


def cleanse_measurement(measurement: str) -> int:
    """Ensure a measurement taken from a spreadsheet is a number.
    The string 'NT' is used to represent 'Not tested' and some cells
    contain '<1' instead of an integer value.

    :param measurement: The value of a measurement from the spreadsheet.
    :return: The measurement as an integer.
    """
    if isinstance(measurement, int):
        return measurement
    elif measurement.startswith(('<', '>')):
        return int(measurement.replace('<', '').replace('>', '').replace(',', ''))
    else:
        return -1


def get_dates(sheet: Worksheet) -> List[datetime]:
    """Extract the measurement dates from the spreadsheet.

    :param sheet: The spreadsheet to read the dates from.
    :return: A list of datetime objects.
    """
    start = f'{FIRST_DATE_COL}{DATE_ROW}'
    end = f'{get_column_letter(sheet.max_column)}{DATE_ROW}'
    dates = [d.value for d in sheet[start:end][0] if d.value]
    return dates


def get_locations(sheet: Worksheet) -> List[str]:
    """Extract the locations from the spreadsheet.

    :param sheet: The spreadsheet to read the locations from.
    :return: A list of location names.
    """
    start = f'{LOCATION_COL}{FIRST_LOCATION_ROW}'
    end = f'{LOCATION_COL}{LAST_LOCATION_ROW}'
    return [l.value.lower().replace(" ", "-") for l, in sheet[start:end]]


def get_measurements(sheet: Worksheet) -> List[Tuple[int, ...]]:
    """Extract the measurements from the spreadsheet.

    :param sheet: The spreadsheet to read the locations from.
    :return: A list of tuples, with each tuple containing measurements for a date.
    """
    col_iter = sheet.iter_cols(
        min_row=FIRST_LOCATION_ROW,
        max_row=LAST_LOCATION_ROW,
        min_col=column_index_from_string(FIRST_DATE_COL),
        max_col=sheet.max_column,
        values_only=True)

    measurements = []

    for col in col_iter:
        if not col[0]:
            continue

        measurements.append(tuple(cleanse_measurement(value) for value in col))

    return measurements


def read_sheet(sheet: Worksheet) -> Dict[str, Dict[str, int]]:
    """Read date, location, and measurement data from a spreadsheet and represent
    this information as a dictionary. Date keys reference dictionaries of location
    and measurement key-value pairs.

    :param sheet: The spreadsheet to read the locations from.
    :return: A nested dictionary representing the dates and locations measurements
     were taken.
    """
    dates = get_dates(sheet)
    locations = get_locations(sheet)
    measurements = get_measurements(sheet)

    results = defaultdict(dict)

    for date_index, date_value in enumerate(dates):
        date_str = date_value.strftime("%Y-%m-%d")

        for location_index, location in enumerate(locations):
            results[date_str][location] = measurements[date_index][location_index]

    return results


if __name__ == "__main__":
    wb = load_workbook(filename=XLSX_FILE)
    result_sheets = [sheet for sheet in wb if 'results' in sheet.title]

    all_results = {}

    for sheet in result_sheets:
        results = read_sheet(sheet)
        all_results.update(results)

    with open(RESULTS_FILE, 'w') as f:
        f.write(json.dumps(all_results, sort_keys=True))
